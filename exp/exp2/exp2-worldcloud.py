import re
import jieba
from os import path
import os
from PIL import Image
from jieba.analyse import *
from openpyxl import load_workbook
from wordcloud import WordCloud, ImageColorGenerator
import wordcloud
import matplotlib.pyplot as plt
import numpy as np
import matplotlib.pyplot as plt
import multidict
from imageio import imread

font_path = r'exp/exp2/share/SourceHanSerifK-Light.otf'

#读取excel文件工作表Sheet1
workbook = load_workbook(r'exp/exp2/share/职位.xlsx')
worksheet = workbook['Sheet1']
#将Sheet1中H列内容写入txt文件
data = []
f = open(r'exp/exp2/output/职位信息.txt', 'w')
for row in range(2, worksheet.max_row + 1):
    data.append(worksheet['H'+str(row)].value)
f.write(str(data))
f.close()

#去除html标签
f = open(r'exp/exp2/output/职位信息.txt', 'r').read()
g = re.sub('<.*?>', '', f)
g = re.sub('nbsp', '', g)
g = re.sub('  ', '', g)

#清洗后数据写入新的txt
h = open(r'exp/exp2/output/职位信息-清洗后.txt', 'w')
h.write(g)
h.close()

#生成词云
#读取txt
f = open(r'exp/exp2/output/职位信息-清洗后.txt', 'r', encoding='UTF8').read()
# f = open(r'exp/exp2/output/职位信息-清洗后.txt', 'r').read()
#分词
sep_list = jieba.lcut(f)
print('分词', type(sep_list), len(sep_list), sep_list)
#过滤停用词
# stopwords为停用词list，stop.txt停用词一行一个
stopwords = [line.strip() for line in open(r'exp/exp2/share/stop.txt', 'r', encoding='utf-8').readlines()]
print('停用词', type(stopwords), stopwords)
outstr = ''
for word in sep_list:
    if word not in stopwords:
        outstr += word
print('过滤停用词后', type(outstr), outstr)
#过滤后重新分词
outstr = jieba.lcut(outstr)
outstr = ' '.join(outstr)
print('再次分词后', type(outstr), outstr)

    
imagename = r'exp/exp2/share/ar15.jpg'
back_coloring = imread(imagename)


wc = WordCloud(font_path=font_path, background_color="white", max_words=2000,
 mask=back_coloring, max_font_size=100, random_state=42, 
 width=1000, height=860, margin=2)

wc.generate(outstr)

image_colors_byImg = ImageColorGenerator(back_coloring)

# show
# we could also give color_func=image_colors directly in the constructor
plt.imshow(wc.recolor(color_func=image_colors_byImg), interpolation="bilinear")
plt.axis("off")
plt.figure()
plt.imshow(back_coloring, interpolation="bilinear")
plt.axis("off")
plt.show()

# save wordcloud
wc.to_file(r'exp/exp2/output/wordcloud.png')

#生成词频
for keyword, weight in extract_tags(outstr, topK=50, withWeight=True, allowPOS=()):
    print('%s %s' % (keyword, weight))
